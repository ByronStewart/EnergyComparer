"""
Energy Made Easy Enhanced Scraper
==================================
Enhanced version that supports distributor selection for postcodes served
by multiple electricity distributors (e.g. boundary postcodes like 2850).

When a postcode has multiple distributors, the scraper will:
  1. Query the meta API to discover available distributors
  2. Test which distributors actually have plans available
  3. Let you choose which distributor's plans to fetch
  4. Optionally fetch plans for ALL distributors at once

For single-distributor postcodes, it works identically to the basic scraper.

Plan Filtering:
  The Energy Made Easy API returns ALL plans for a postcode, including plans
  that require special metering (demand tariffs) or controlled load circuits.
  By default, this scraper filters plans to match what the website shows:
    - Plans with demand charges are excluded (require demand meters)
    - Plans with controlled load are excluded unless you specify --controlled-load
  Use --no-filter to get ALL plans unfiltered.

NMI (National Meter Identifier) Note:
  The Energy Made Easy NMI API requires browser-session authentication
  that cannot be replicated via simple HTTP requests (returns 403).
  NMI-based personalised results are not supported at this time.

Usage:
    python scraper_enhanced.py                           # Interactive mode
    python scraper_enhanced.py 2850                      # Multi-distributor postcode
    python scraper_enhanced.py 2850 --dist 13            # Specify distributor ID directly
    python scraper_enhanced.py 2850 --dist all           # Fetch all distributors
    python scraper_enhanced.py 2000                      # Single-distributor (auto)
    python scraper_enhanced.py 2000 --fuel gas           # Gas plans
    python scraper_enhanced.py 2000 --type business      # Business plans
    python scraper_enhanced.py 4075 --controlled-load    # Include controlled load plans
    python scraper_enhanced.py 4075 --no-filter          # Get ALL plans (no filtering)
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

API_BASE = "https://api.energymadeeasy.gov.au"
POSTCODE_API = f"{API_BASE}/location/postcodes"
PLANS_API = f"{API_BASE}/consumerplan/plans"
META_API = f"{API_BASE}/consumerplan/plans"  # /{postcode}/meta?fuelType=E

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/144.0.0.0 Safari/537.36"
    ),
    "Accept": "*/*",
    "Origin": "https://www.energymadeeasy.gov.au",
    "Referer": "https://www.energymadeeasy.gov.au/",
}

PAYMENT_OPTIONS = {
    "P": "Post/Mail",
    "DD": "Direct Debit",
    "CC": "Credit Card",
    "BP": "BPay",
    "CP": "Centrepay",
}

GST_MULTIPLIER = 1.1  # 10% GST - API returns ex-GST values, website shows inc-GST

FEE_TYPES = {
    "ConnF": "Connection Fee",
    "DiscoF": "Disconnection Fee",
    "DiscoFMO": "Disconnection Fee (Move Out)",
    "DiscoFNP": "Disconnection Fee (Non-Payment)",
    "ChDF": "Charge Dispute Fee",
    "DDF": "Dishonoured Direct Debit Fee",
    "LPF": "Late Payment Fee",
    "PBF": "Paper Bill Fee",
    "CCF": "Credit Card Fee",
    "PPF": "Payment Processing Fee",
    "RecoF": "Reconnection Fee",
    "MBSF": "Membership Fee",
    "OF": "Other Fee",
}


# ---------------------------------------------------------------------------
# API Functions
# ---------------------------------------------------------------------------


def validate_postcode(postcode: str) -> list[dict]:
    """Validate a postcode and return matching locations."""
    url = f"{POSTCODE_API}/{postcode}"
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    locations = data.get("data", [])
    return locations


def fetch_distributors(postcode: str, fuel_type: str = "E") -> list[dict]:
    """Fetch available electricity distributors for a postcode from the meta API.

    Returns a list of dicts with 'id' and 'name' keys, deduplicated.
    """
    url = f"{META_API}/{postcode}/meta?fuelType={fuel_type}"
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    seen = set()
    distributors = []
    for item in data.get("data", []):
        for sa in item.get("planData", {}).get("supplyArea", []):
            key = sa["id"]
            if key not in seen:
                seen.add(key)
                distributors.append({"id": sa["id"], "name": sa["name"]})

    return sorted(distributors, key=lambda d: d["name"])


def probe_distributor_plans(postcode: str, dist_id: str, fuel_type: str = "E",
                            customer_type: str = "R") -> int:
    """Quick probe to check how many plans a distributor has. Returns count or -1 on error."""
    params = {
        "usageDataSource": "noUsageFrontier",
        "customerType": customer_type,
        "distE": dist_id if fuel_type == "E" else "",
        "distG": dist_id if fuel_type == "G" else "",
        "fuelType": fuel_type,
        "journey": fuel_type,
        "postcode": postcode,
    }
    try:
        resp = requests.get(PLANS_API, params=params, headers=HEADERS, timeout=60)
        if resp.status_code != 200:
            return -1
        plans = resp.json().get("data", {}).get("plans", [])
        return len(plans)
    except Exception:
        return -1


def fetch_plans(postcode: str, fuel_type: str = "E", customer_type: str = "R",
                dist_id: str = "") -> list[dict]:
    """Fetch all energy plans for a given postcode and optional distributor."""
    dist_e = dist_id if fuel_type == "E" else ""
    dist_g = dist_id if fuel_type == "G" else ""
    params = {
        "usageDataSource": "noUsageFrontier",
        "customerType": customer_type,
        "distE": dist_e,
        "distG": dist_g,
        "fuelType": fuel_type,
        "journey": fuel_type,
        "postcode": postcode,
    }
    resp = requests.get(PLANS_API, params=params, headers=HEADERS, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    plans = data.get("data", {}).get("plans", [])
    return plans


# ---------------------------------------------------------------------------
# Data Extraction Helpers
# ---------------------------------------------------------------------------


def extract_supply_charge(contract: dict) -> float | None:
    """Extract daily supply charge in c/day (inc. GST). Returns numeric value."""
    tariff_periods = contract.get("tariffPeriod", [])
    if tariff_periods:
        charge = tariff_periods[0].get("dailySupplyCharge")
        if charge is not None:
            return round(charge * GST_MULTIPLIER, 2)
    return None


def _collect_usage_rates(contract: dict) -> list[float]:
    """Collect all usage rates from a contract (inc. GST)."""
    tariff_periods = contract.get("tariffPeriod", [])
    if not tariff_periods:
        return []
    tp = tariff_periods[0]
    pricing_model = contract.get("pricingModel", "SR")

    if pricing_model == "TOU":
        rates = []
        for block in tp.get("touBlock", []):
            for br in block.get("blockRate", []):
                rates.append(round(br["unitPrice"] * GST_MULTIPLIER, 2))
        return rates
    else:
        return [round(br["unitPrice"] * GST_MULTIPLIER, 2) for br in tp.get("blockRate", [])]


def extract_usage_rate_min(contract: dict) -> float | None:
    """Extract the lowest usage rate (inc. GST)."""
    rates = _collect_usage_rates(contract)
    return min(rates) if rates else None


def extract_usage_rate_max(contract: dict) -> float | None:
    """Extract the highest usage rate (inc. GST)."""
    rates = _collect_usage_rates(contract)
    return max(rates) if rates else None


def extract_tou_peak_rate(contract: dict) -> float | None:
    """Extract peak rate for TOU plans (inc. GST)."""
    if contract.get("pricingModel") != "TOU":
        return None
    rates = _collect_usage_rates(contract)
    return max(rates) if rates else None


def extract_tou_offpeak_rate(contract: dict) -> float | None:
    """Extract off-peak rate for TOU plans (inc. GST)."""
    if contract.get("pricingModel") != "TOU":
        return None
    rates = _collect_usage_rates(contract)
    return min(rates) if rates else None


def _collect_solar_fit_rates(contract: dict) -> list[float]:
    """Collect all non-zero *retailer* solar feed-in tariff rates.

    The API returns two kinds of solar FIT entries distinguished by ``type``:
      - ``"R"`` (Retailer) – the actual feed-in tariff offered by the plan.
      - ``"G"`` (Government) – legacy government bonus schemes (e.g. the
        Queensland Solar Bonus Scheme at 44 c/kWh).  These are paid by the
        government to grandfathered customers and are **not** part of the
        retailer's offer.  Including them inflates the reported FIT and
        misleads users who are not on the legacy scheme.

    Only retailer (``type != "G"``) entries are collected.
    """
    solar_fits = contract.get("solarFit", [])
    if not solar_fits:
        return []
    rates = []
    for fit in solar_fits:
        # Skip government / legacy bonus scheme entries
        if fit.get("type") == "G":
            continue
        rate = fit.get("rate")
        if rate is not None:
            rates.append(rate)
        else:
            for sr in fit.get("singleTariffRates", []):
                rates.append(sr.get("unitPrice", 0))
    return [r for r in rates if r > 0]


def extract_solar_fit_min(contract: dict) -> float:
    """Extract the lowest solar feed-in tariff rate (c/kWh, GST exempt)."""
    non_zero = _collect_solar_fit_rates(contract)
    return round(min(non_zero), 2) if non_zero else 0.0


def extract_solar_fit_max(contract: dict) -> float:
    """Extract the highest solar feed-in tariff rate (c/kWh, GST exempt)."""
    non_zero = _collect_solar_fit_rates(contract)
    return round(max(non_zero), 2) if non_zero else 0.0


def extract_solar_fit_details(contract: dict) -> str:
    """Extract detailed solar FIT info including volumes/tiers.

    Only retailer FIT entries (``type != "G"``) are included.
    Government/legacy bonus scheme entries are excluded.
    """
    solar_fits = contract.get("solarFit", [])
    if not solar_fits:
        return "No solar feed-in tariff"
    details = []
    for fit in solar_fits:
        # Skip government / legacy bonus scheme entries
        if fit.get("type") == "G":
            continue
        single_rates = fit.get("singleTariffRates", [])
        for sr in single_rates:
            price = sr.get("unitPrice", 0)
            volume = sr.get("volume", 0)
            if price > 0:
                if volume > 0:
                    details.append(f"{price}c/kWh (first {volume}kWh/day)")
                else:
                    details.append(f"{price}c/kWh")
    if not details:
        return "No solar feed-in tariff"
    return "; ".join(details)


def extract_controlled_load(contract: dict) -> str:
    """Extract controlled load rates if present (inc. GST)."""
    cl_list = contract.get("controlledLoad", [])
    if not cl_list:
        return "N/A"
    parts = []
    for cl in cl_list:
        sr = cl.get("singleRate", {})
        name = sr.get("displayName", "Controlled Load")
        rates_list = sr.get("rates", [])
        daily_charge = sr.get("dailySupplyCharge", 0)
        rate_strs = [f"{r['unitPrice'] * GST_MULTIPLIER:.2f}c/kWh" for r in rates_list]
        part = f"{name}: {', '.join(rate_strs)}"
        if daily_charge:
            part += f" + {daily_charge * GST_MULTIPLIER:.2f}c/day supply"
        parts.append(part)
    return "; ".join(parts) if parts else "N/A"


def extract_controlled_load_rate(contract: dict) -> float:
    """Extract the primary controlled load usage rate in c/kWh (inc. GST).

    If multiple controlled load entries exist, returns the rate from the first one.
    Returns 0.0 if no controlled load is present.
    """
    cl_list = contract.get("controlledLoad", [])
    if not cl_list:
        return 0.0
    sr = cl_list[0].get("singleRate", {})
    rates_list = sr.get("rates", [])
    if rates_list:
        return round(rates_list[0].get("unitPrice", 0) * GST_MULTIPLIER, 2)
    return 0.0


def extract_controlled_load_supply(contract: dict) -> float:
    """Extract the controlled load daily supply charge in c/day (inc. GST).

    Returns 0.0 if no controlled load or no supply charge.
    """
    cl_list = contract.get("controlledLoad", [])
    if not cl_list:
        return 0.0
    sr = cl_list[0].get("singleRate", {})
    daily_charge = sr.get("dailySupplyCharge", 0)
    return round(daily_charge * GST_MULTIPLIER, 2) if daily_charge else 0.0


def extract_discounts(contract: dict) -> str:
    """Extract discount information."""
    discounts = contract.get("discount", [])
    if not discounts:
        return "None"
    parts = []
    for d in discounts:
        name = d.get("name", "Discount")
        pct = d.get("discountPercent")
        amt = d.get("discountAmount")
        if pct:
            parts.append(f"{name} ({pct}%)")
        elif amt:
            parts.append(f"{name} (${amt})")
        else:
            parts.append(name)
    return "; ".join(parts)


def extract_fees(contract: dict) -> str:
    """Extract fees as a formatted string."""
    fees = contract.get("fee", [])
    if not fees:
        return "None"
    parts = []
    for f in fees:
        fee_name = FEE_TYPES.get(f.get("feeType", ""), f.get("feeType", "Unknown"))
        amount = f.get("amount", 0)
        parts.append(f"{fee_name}: ${amount:.2f}")
    return "; ".join(parts)


def extract_payment_options(contract: dict) -> str:
    """Extract payment options."""
    options = contract.get("paymentOption", [])
    if not options:
        return "N/A"
    return ", ".join(PAYMENT_OPTIONS.get(o, o) for o in options)


def extract_meter_types(contract: dict) -> str:
    """Extract supported meter types."""
    meters = contract.get("meterType", [])
    if not meters:
        return "N/A"
    meter_map = {
        "Type 6": "Basic Meter",
        "Type 4": "Smart Meter",
        "Type 4a": "Smart Meter (4a)",
        "Type 1": "Interval Meter",
    }
    readable = []
    for m in meters:
        readable.append(meter_map.get(m, m))
    seen = set()
    unique = []
    for r in readable:
        if r not in seen:
            seen.add(r)
            unique.append(r)
    return ", ".join(unique)


def extract_benefit_period(contract: dict) -> str:
    """Extract benefit period."""
    bp = contract.get("benefitPeriod", "")
    if not bp:
        return "N/A"
    return bp


def extract_contract_term(contract: dict) -> str:
    """Extract contract term type."""
    term = contract.get("termType", "")
    term_map = {
        "E": "No lock-in",
        "1": "1 year",
        "2": "2 years",
        "3": "3 years",
    }
    return term_map.get(term, term if term else "N/A")


# ---------------------------------------------------------------------------
# Plan Filtering
# ---------------------------------------------------------------------------


def plan_has_demand_charge(plan: dict) -> bool:
    """Check if a plan has demand charges in any tariff period.

    Plans with demand charges require a demand meter/tariff which most
    standard residential customers do not have.
    """
    contract = plan["planData"]["contract"][0]
    for tp in contract.get("tariffPeriod", []):
        dc = tp.get("demandCharge")
        if dc and len(dc) > 0:
            return True
    return False


def plan_has_controlled_load(plan: dict) -> bool:
    """Check if a plan includes controlled load pricing.

    Controlled load plans are for customers with a separate circuit for
    hot water, pool pumps, or floor heating on a controlled load tariff.
    """
    contract = plan["planData"]["contract"][0]
    cl = contract.get("controlledLoad")
    return bool(cl and len(cl) > 0)


def filter_plans(
    plans: list[dict],
    include_controlled_load: bool = False,
    include_demand: bool = False,
) -> tuple[list[dict], dict]:
    """Filter plans to match what a standard residential customer can use.

    The Energy Made Easy website applies the same filters by default:
      - Plans with demand charges are excluded (require demand meters)
      - Plans with controlled load are excluded (require CL circuit)

    Args:
        plans: Raw plan list from the API.
        include_controlled_load: If True, include plans with controlled load.
        include_demand: If True, include plans with demand charges.

    Returns:
        Tuple of (filtered_plans, stats_dict) where stats_dict contains
        counts of what was filtered out.
    """
    stats = {
        "total": len(plans),
        "demand_filtered": 0,
        "controlled_load_filtered": 0,
        "kept": 0,
    }

    filtered = []
    for plan in plans:
        if not include_demand and plan_has_demand_charge(plan):
            stats["demand_filtered"] += 1
            continue
        if not include_controlled_load and plan_has_controlled_load(plan):
            stats["controlled_load_filtered"] += 1
            continue
        filtered.append(plan)

    stats["kept"] = len(filtered)
    return filtered, stats


# ---------------------------------------------------------------------------
# Plan Processing
# ---------------------------------------------------------------------------


def build_plan_url(plan_id: str, postcode: str) -> str:
    """Build the Energy Made Easy plan detail URL."""
    return (
        f"https://www.energymadeeasy.gov.au/plan"
        f"?id={plan_id}&postcode={postcode}"
        f"&pricingPeriod=yearly&withDiscounts=true&benchmarkUsage=medium"
    )


def process_plan(plan: dict, postcode: str, distributor_name: str = "") -> dict:
    """Extract all relevant fields from a plan into a flat dictionary."""
    plan_data = plan["planData"]
    contract = plan_data["contract"][0]  # Primary contract
    pcr = plan.get("pcr", {}).get("costs", {})

    # Get estimated costs for different usage levels
    fuel_key = "electricity" if plan_data["fuelType"] == "E" else "gas"
    costs = pcr.get(fuel_key, {})

    plan_id = plan_data.get("planId", "")

    row = {
        "Plan ID": plan_id,
        "Plan Name": plan_data.get("planName", ""),
        "Retailer": plan_data.get("retailerName", ""),
        "Distributor": distributor_name if distributor_name else "N/A",
        "Plan URL": build_plan_url(plan_id, postcode),
        "Tariff Type": plan_data.get("tariffType", ""),
        "Pricing Model": contract.get("pricingModel", ""),
        "Contract Term": extract_contract_term(contract),
        "Benefit Period": extract_benefit_period(contract),
        "Supply Charge (c/day)": extract_supply_charge(contract),
        "Usage Rate Min (c/kWh)": extract_usage_rate_min(contract),
        "Usage Rate Max (c/kWh)": extract_usage_rate_max(contract),
        "Peak Rate (c/kWh)": extract_tou_peak_rate(contract),
        "Off-Peak Rate (c/kWh)": extract_tou_offpeak_rate(contract),
        "Solar FIT Min (c/kWh)": extract_solar_fit_min(contract),
        "Solar FIT Max (c/kWh)": extract_solar_fit_max(contract),
        "Solar FIT Details": extract_solar_fit_details(contract),
        "Controlled Load": extract_controlled_load(contract),
        "CL Rate (c/kWh)": extract_controlled_load_rate(contract),
        "CL Supply (c/day)": extract_controlled_load_supply(contract),
        "Discounts": extract_discounts(contract),
        "Fees": extract_fees(contract),
        "Payment Options": extract_payment_options(contract),
        "Meter Types": extract_meter_types(contract),
        "Est. Cost/Year (Low Usage)": costs.get("small", {}).get("yearly", {}).get("allDiscounts", None),
        "Est. Cost/Year (Medium Usage)": costs.get("medium", {}).get("yearly", {}).get("allDiscounts", None),
        "Est. Cost/Year (High Usage)": costs.get("large", {}).get("yearly", {}).get("allDiscounts", None),
        "Est. Cost/Year (Low, No Disc.)": costs.get("small", {}).get("yearly", {}).get("noDiscounts", None),
        "Est. Cost/Year (Medium, No Disc.)": costs.get("medium", {}).get("yearly", {}).get("noDiscounts", None),
        "Est. Cost/Year (High, No Disc.)": costs.get("large", {}).get("yearly", {}).get("noDiscounts", None),
    }

    return row


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

# Column groupings for coloring
COLUMN_GROUPS = {
    "identity": ["Plan ID", "Plan Name", "Retailer", "Distributor", "Plan URL"],
    "plan_type": ["Tariff Type", "Pricing Model", "Contract Term", "Benefit Period"],
    "charges": ["Supply Charge (c/day)", "Usage Rate Min (c/kWh)", "Usage Rate Max (c/kWh)", "Peak Rate (c/kWh)", "Off-Peak Rate (c/kWh)"],
    "solar": ["Solar FIT Min (c/kWh)", "Solar FIT Max (c/kWh)", "Solar FIT Details"],
    "extras": ["Controlled Load", "CL Rate (c/kWh)", "CL Supply (c/day)", "Discounts", "Fees", "Payment Options", "Meter Types"],
    "cost_disc": ["Est. Cost/Year (Low Usage)", "Est. Cost/Year (Medium Usage)", "Est. Cost/Year (High Usage)"],
    "cost_no_disc": ["Est. Cost/Year (Low, No Disc.)", "Est. Cost/Year (Medium, No Disc.)", "Est. Cost/Year (High, No Disc.)"],
}

# Columns that hold numeric values and should be formatted as numbers in Excel
NUMERIC_COLUMNS = {
    "Supply Charge (c/day)",
    "Usage Rate Min (c/kWh)",
    "Usage Rate Max (c/kWh)",
    "Peak Rate (c/kWh)",
    "Off-Peak Rate (c/kWh)",
    "Solar FIT Min (c/kWh)",
    "Solar FIT Max (c/kWh)",
    "CL Rate (c/kWh)",
    "CL Supply (c/day)",
    "Est. Cost/Year (Low Usage)",
    "Est. Cost/Year (Medium Usage)",
    "Est. Cost/Year (High Usage)",
    "Est. Cost/Year (Low, No Disc.)",
    "Est. Cost/Year (Medium, No Disc.)",
    "Est. Cost/Year (High, No Disc.)",
}

# Columns that should be rendered as clickable hyperlinks
LINK_COLUMNS = {"Plan URL"}

GROUP_COLORS = {
    "identity": "D6EAF8",
    "plan_type": "D5F5E3",
    "charges": "FDEBD0",
    "solar": "F9E79F",
    "extras": "E8DAEF",
    "cost_disc": "D4EFDF",
    "cost_no_disc": "FADBD8",
}


def get_column_group(col_name: str) -> str | None:
    """Get the group a column belongs to."""
    for group, cols in COLUMN_GROUPS.items():
        if col_name in cols:
            return group
    return None


def export_to_excel(plans_data: list[dict], postcode: str, fuel_type: str,
                    customer_type: str, distributor_info: str = "") -> str:
    """Export plan data to a formatted Excel spreadsheet."""
    wb = Workbook()

    # ---- Summary sheet ----
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title
    ws_summary["A1"] = "Energy Made Easy - Plan Comparison (Enhanced)"
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=16, color="1B4F72")
    ws_summary.merge_cells("A1:D1")

    ws_summary["A3"] = "Postcode:"
    ws_summary["B3"] = postcode
    ws_summary["A4"] = "Fuel Type:"
    ws_summary["B4"] = "Electricity" if fuel_type == "E" else "Gas"
    ws_summary["A5"] = "Customer Type:"
    ws_summary["B5"] = "Residential" if customer_type == "R" else "Small Business"
    ws_summary["A6"] = "Distributor:"
    ws_summary["B6"] = distributor_info if distributor_info else "All / Auto"
    ws_summary["A7"] = "Total Plans Found:"
    ws_summary["B7"] = len(plans_data)
    ws_summary["A8"] = "Date Scraped:"
    ws_summary["B8"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary["A9"] = "Source:"
    ws_summary["B9"] = "https://www.energymadeeasy.gov.au/"

    for row in range(3, 10):
        ws_summary[f"A{row}"].font = Font(name="Calibri", bold=True, size=11)
        ws_summary[f"B{row}"].font = Font(name="Calibri", size=11)

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 50

    # Count by retailer
    ws_summary["A11"] = "Plans by Retailer"
    ws_summary["A11"].font = Font(name="Calibri", bold=True, size=13, color="1B4F72")
    ws_summary.merge_cells("A11:B11")

    retailer_counts = {}
    for p in plans_data:
        r = p["Retailer"]
        retailer_counts[r] = retailer_counts.get(r, 0) + 1

    row_idx = 12
    ws_summary[f"A{row_idx}"] = "Retailer"
    ws_summary[f"B{row_idx}"] = "Number of Plans"
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws_summary[f"A{row_idx}"].fill = header_fill
    ws_summary[f"B{row_idx}"].fill = header_fill
    ws_summary[f"A{row_idx}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws_summary[f"B{row_idx}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")

    for retailer, count in sorted(retailer_counts.items()):
        row_idx += 1
        ws_summary[f"A{row_idx}"] = retailer
        ws_summary[f"B{row_idx}"] = count
        if row_idx % 2 == 0:
            alt_fill = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
            ws_summary[f"A{row_idx}"].fill = alt_fill
            ws_summary[f"B{row_idx}"].fill = alt_fill

    # Count by distributor (if multiple)
    dist_counts = {}
    for p in plans_data:
        d = p.get("Distributor", "N/A")
        dist_counts[d] = dist_counts.get(d, 0) + 1

    if len(dist_counts) > 1:
        row_idx += 2
        ws_summary[f"A{row_idx}"] = "Plans by Distributor"
        ws_summary[f"A{row_idx}"].font = Font(name="Calibri", bold=True, size=13, color="1B4F72")
        ws_summary.merge_cells(f"A{row_idx}:B{row_idx}")
        row_idx += 1
        ws_summary[f"A{row_idx}"] = "Distributor"
        ws_summary[f"B{row_idx}"] = "Number of Plans"
        ws_summary[f"A{row_idx}"].fill = header_fill
        ws_summary[f"B{row_idx}"].fill = header_fill
        ws_summary[f"A{row_idx}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        ws_summary[f"B{row_idx}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        for dist_name, count in sorted(dist_counts.items()):
            row_idx += 1
            ws_summary[f"A{row_idx}"] = dist_name
            ws_summary[f"B{row_idx}"] = count
            if row_idx % 2 == 0:
                alt_fill = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
                ws_summary[f"A{row_idx}"].fill = alt_fill
                ws_summary[f"B{row_idx}"].fill = alt_fill

    # ---- All Plans sheet ----
    ws_all = wb.create_sheet("All Plans")
    _write_plans_sheet(ws_all, plans_data)

    # ---- Single Rate sheet ----
    sr_plans = [p for p in plans_data if p["Pricing Model"] == "SR"]
    if sr_plans:
        ws_sr = wb.create_sheet("Single Rate Plans")
        _write_plans_sheet(ws_sr, sr_plans)

    # ---- Time of Use sheet ----
    tou_plans = [p for p in plans_data if p["Pricing Model"] == "TOU"]
    if tou_plans:
        ws_tou = wb.create_sheet("Time of Use Plans")
        _write_plans_sheet(ws_tou, tou_plans)

    # ---- Solar-Friendly sheet (sorted by highest FIT) ----
    solar_plans = [p for p in plans_data if (p.get("Solar FIT Max (c/kWh)") or 0) > 0]
    if solar_plans:
        solar_plans.sort(key=lambda p: p.get("Solar FIT Max (c/kWh)") or 0, reverse=True)
        ws_solar = wb.create_sheet("Best Solar FIT")
        _write_plans_sheet(ws_solar, solar_plans)

    # ---- Cheapest Plans sheet ----
    cheap_plans = [p for p in plans_data if p.get("Est. Cost/Year (Medium Usage)") is not None]
    if cheap_plans:
        cheap_plans.sort(key=lambda p: p["Est. Cost/Year (Medium Usage)"])
        ws_cheap = wb.create_sheet("Cheapest Plans")
        _write_plans_sheet(ws_cheap, cheap_plans[:50])  # Top 50

    # ---- Plan Calculator sheet ----
    # Only include plans that have enough rate data for meaningful calculations
    calc_plans = [p for p in plans_data
                  if (p.get("Supply Charge (c/day)") is not None
                      and ((p.get("Usage Rate Max (c/kWh)") is not None and p["Pricing Model"] == "SR")
                           or (p.get("Peak Rate (c/kWh)") is not None and p["Pricing Model"] == "TOU")))]
    calc_info = None
    if calc_plans:
        ws_calc = wb.create_sheet("Plan Calculator")
        calc_info = _write_calculator_sheet(ws_calc, calc_plans)

    # Save
    dist_suffix = ""
    if distributor_info and distributor_info not in ("All / Auto", "Auto"):
        # Sanitise distributor name for filename
        safe_name = "".join(c if c.isalnum() or c in (" ", "-") else "_" for c in distributor_info)
        safe_name = safe_name.strip().replace(" ", "_")[:30]
        dist_suffix = f"_{safe_name}"

    filename = f"energy_plans_{postcode}_{fuel_type}{dist_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = Path(filename)
    wb.save(filepath)

    # Inject VBA sort macro and convert to .xlsm if Plan Calculator was created
    if calc_info is not None:
        header_row, plan_count = calc_info
        try:
            xlsm_path = _inject_vba_and_save_as_xlsm(str(filepath), header_row, plan_count)
            return xlsm_path
        except Exception as e:
            print(f"  Warning: Could not inject VBA macro ({e}). Saving as .xlsx instead.")
            return str(filepath)

    return str(filepath)


def _write_plans_sheet(ws, plans_data: list[dict]):
    """Write plan data to a worksheet with formatting."""
    if not plans_data:
        return

    columns = list(plans_data[0].keys())
    thin_border = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )

    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    # Write headers
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Write data rows
    data_font = Font(name="Calibri", size=10)
    for row_idx, plan in enumerate(plans_data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = plan.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx)

            # Handle hyperlink columns
            if col_name in LINK_COLUMNS and value:
                cell.hyperlink = value
                cell.value = "View Plan"
                cell.font = link_font
            else:
                cell.value = value
                cell.font = data_font

            # Apply number format for numeric columns
            if col_name in NUMERIC_COLUMNS and isinstance(value, (int, float)):
                if "Est. Cost" in col_name:
                    cell.number_format = "$#,##0"
                else:
                    cell.number_format = "0.00"

            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

            # Apply group coloring
            group = get_column_group(col_name)
            if group and row_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color=GROUP_COLORS[group],
                    end_color=GROUP_COLORS[group],
                    fill_type="solid",
                )
            elif row_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color="F8F9F9",
                    end_color="F8F9F9",
                    fill_type="solid",
                )

    # Auto-fit column widths (with max cap)
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name) + 2
        for row_idx in range(2, min(len(plans_data) + 2, 52)):  # Sample first 50 rows
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, min(len(cell_value), 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(plans_data) + 1}"


# ---------------------------------------------------------------------------
# Usage Profile Presets for TOU Calculator
# ---------------------------------------------------------------------------

# Each preset defines the percentage of daily usage that falls in the peak
# period vs. the off-peak period.  These are rough heuristics since actual
# TOU schedules vary by distributor, but they let the user model different
# consumption patterns.

USAGE_PROFILES = {
    "Flat Usage":         {"peak_pct": 0.50, "offpeak_pct": 0.50,
                           "description": "Even usage across peak and off-peak (50/50)"},
    "Slight Peak":        {"peak_pct": 0.60, "offpeak_pct": 0.40,
                           "description": "Slightly more usage during peak hours (60/40)"},
    "Heavy Peak":         {"peak_pct": 0.75, "offpeak_pct": 0.25,
                           "description": "Most usage during peak hours (75/25)"},
    "Off-Peak Heavy":     {"peak_pct": 0.30, "offpeak_pct": 0.70,
                           "description": "Most usage shifted to off-peak (30/70)"},
    "Battery Optimised":  {"peak_pct": 0.10, "offpeak_pct": 0.90,
                           "description": "Battery covers peak; almost all off-peak (10/90)"},
}

# The profile names in the order they appear in the dropdown
PROFILE_NAMES = list(USAGE_PROFILES.keys())


def _parse_solar_fit_tiers(plan: dict) -> list[tuple[float, float]]:
    """Parse solar FIT tiers from a processed plan dict.

    Returns a list of ``(rate, volume)`` tuples sorted by volume descending
    (highest-volume tier first).  A volume of ``0`` means "all remaining".

    Example return for "10c/kWh (first 8kWh/day); 3c/kWh":
        [(10.0, 8.0), (3.0, 0.0)]
    """
    details = plan.get("Solar FIT Details", "")
    if not details or details == "No solar feed-in tariff":
        return []

    tiers = []
    # Each tier is separated by "; "
    for part in details.split("; "):
        # Match patterns like "10c/kWh (first 8kWh/day)" or "3c/kWh"
        m = re.match(r"([\d.]+)c/kWh(?:\s*\(first\s+([\d.]+)kWh/day\))?", part.strip())
        if m:
            rate = float(m.group(1))
            volume = float(m.group(2)) if m.group(2) else 0.0
            tiers.append((rate, volume))

    # Sort: tiers with a volume cap first, then the uncapped remainder tier
    tiers.sort(key=lambda t: (t[1] == 0, -t[1]))
    return tiers


def _write_calculator_sheet(ws, plans_data: list[dict]) -> tuple[int, int]:
    """Write the Plan Calculator sheet with user inputs and per-plan cost formulas.

    Layout:
        Rows 1-8:   Input section (user-editable cells highlighted in yellow)
                     Includes daily usage, solar export, TOU profile,
                     controlled load toggle (Yes/No), and CL daily kWh.
        Row  9:     Blank separator
        Rows 10-15: Profile legend
        Row  16:    Column headers for the plan comparison table
        Row  17+:   One row per plan with Excel formulas referencing the inputs

    The user can change the input cells and all costs recalculate automatically.

    Controlled load is toggled via a Yes/No dropdown in B7.  When "Yes",
    the CL Cost/day column uses B8 (CL kWh/day) * plan CL rate + CL supply.
    When "No", CL Cost/day is zero and has no effect on the net cost.

    Solar FIT tiers are modelled properly.  For example, a plan with
    "10c/kWh first 8kWh/day; 3c/kWh thereafter" will calculate the credit as:
        MIN(export, 8) * 10  +  MAX(export - 8, 0) * 3

    Returns:
        Tuple of (header_row, plan_count) for use by the VBA macro injector.
    """
    if not plans_data:
        return

    thin_border = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )
    label_font = Font(name="Calibri", bold=True, size=11, color="1B4F72")
    input_font = Font(name="Calibri", size=11)
    input_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")  # pale yellow
    info_font = Font(name="Calibri", size=10, italic=True, color="7F8C8D")

    # ---- Title ----
    ws["A1"] = "Plan Calculator"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1B4F72")
    ws.merge_cells("A1:F1")

    ws["A2"] = "Enter your estimated usage below. All costs update automatically."
    ws["A2"].font = info_font
    ws.merge_cells("A2:F2")

    # ---- User Inputs ----
    # Row 4: Daily usage
    ws["A4"] = "Daily Usage (kWh):"
    ws["A4"].font = label_font
    ws["B4"] = 20  # default 20 kWh/day
    ws["B4"].font = input_font
    ws["B4"].fill = input_fill
    ws["B4"].border = thin_border
    ws["B4"].number_format = "0.0"
    ws["C4"] = "Your estimated daily electricity consumption"
    ws["C4"].font = info_font

    # Row 5: Solar export
    ws["A5"] = "Daily Solar Export (kWh):"
    ws["A5"].font = label_font
    ws["B5"] = 10  # default 10 kWh/day export
    ws["B5"].font = input_font
    ws["B5"].fill = input_fill
    ws["B5"].border = thin_border
    ws["B5"].number_format = "0.0"
    ws["C5"] = "How much solar you expect to export back to the grid per day"
    ws["C5"].font = info_font

    # Row 6: Usage profile (dropdown)
    ws["A6"] = "Usage Profile (TOU plans):"
    ws["A6"].font = label_font
    ws["B6"] = PROFILE_NAMES[0]  # default "Flat Usage"
    ws["B6"].font = input_font
    ws["B6"].fill = input_fill
    ws["B6"].border = thin_border

    # Add data validation dropdown for usage profile
    profile_list = ",".join(PROFILE_NAMES)
    dv = DataValidation(
        type="list",
        formula1=f'"{profile_list}"',
        allow_blank=False,
    )
    dv.error = "Please select a usage profile from the list."
    dv.errorTitle = "Invalid Profile"
    dv.prompt = "Select how your usage is distributed across peak/off-peak periods."
    dv.promptTitle = "Usage Profile"
    ws.add_data_validation(dv)
    dv.add(ws["B6"])

    ws["C6"] = "Controls peak/off-peak split for Time-of-Use plans"
    ws["C6"].font = info_font

    # Row 7: Controlled Load (Yes/No dropdown)
    ws["A7"] = "Controlled Load:"
    ws["A7"].font = label_font
    ws["B7"] = "No"
    ws["B7"].font = input_font
    ws["B7"].fill = input_fill
    ws["B7"].border = thin_border

    dv_cl = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False,
    )
    dv_cl.error = "Please select Yes or No."
    dv_cl.errorTitle = "Invalid Choice"
    dv_cl.prompt = "Do you have a controlled load circuit (e.g. off-peak hot water)?"
    dv_cl.promptTitle = "Controlled Load"
    ws.add_data_validation(dv_cl)
    dv_cl.add(ws["B7"])

    ws["C7"] = "Select Yes if you have a controlled load circuit (hot water, pool pump, etc.)"
    ws["C7"].font = info_font

    # Row 8: Controlled Load kWh (only relevant when CL = Yes)
    ws["A8"] = "Controlled Load Usage (kWh/day):"
    ws["A8"].font = label_font
    ws["B8"] = 8  # default 8 kWh/day (typical hot water system)
    ws["B8"].font = input_font
    ws["B8"].fill = input_fill
    ws["B8"].border = thin_border
    ws["B8"].number_format = "0.0"
    ws["C8"] = "Daily kWh on your controlled load circuit (only used when Controlled Load = Yes)"
    ws["C8"].font = info_font

    # Row 10-14: Profile legend
    ws["A10"] = "Profile Reference:"
    ws["A10"].font = Font(name="Calibri", bold=True, size=10, color="1B4F72")
    ws.merge_cells("A10:C10")

    legend_row = 11
    for name, profile in USAGE_PROFILES.items():
        ws[f"A{legend_row}"] = name
        ws[f"A{legend_row}"].font = Font(name="Calibri", bold=True, size=9)
        peak_pct = int(profile["peak_pct"] * 100)
        offpeak_pct = int(profile["offpeak_pct"] * 100)
        ws[f"B{legend_row}"] = f"Peak {peak_pct}% / Off-Peak {offpeak_pct}%"
        ws[f"B{legend_row}"].font = Font(name="Calibri", size=9)
        ws[f"C{legend_row}"] = profile["description"]
        ws[f"C{legend_row}"].font = Font(name="Calibri", size=9, italic=True, color="7F8C8D")
        legend_row += 1

    # ---- Column widths for input section ----
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 55

    # ---- Build the peak/off-peak percentage lookup using nested IFs ----
    # This formula returns the peak fraction based on the profile selected in B6
    peak_if_parts = []
    for name, profile in USAGE_PROFILES.items():
        peak_if_parts.append(f'IF($B$6="{name}",{profile["peak_pct"]}')
    # Nest them: IF(B6="Flat Usage",0.5,IF(B6="Slight Peak",0.6,...,0.5))
    peak_formula_inner = "0.5"  # fallback
    for part in reversed(peak_if_parts):
        peak_formula_inner = f"{part},{peak_formula_inner})"

    offpeak_if_parts = []
    for name, profile in USAGE_PROFILES.items():
        offpeak_if_parts.append(f'IF($B$6="{name}",{profile["offpeak_pct"]}')
    offpeak_formula_inner = "0.5"
    for part in reversed(offpeak_if_parts):
        offpeak_formula_inner = f"{part},{offpeak_formula_inner})"

    # ---- Plan Comparison Table ----
    header_row = legend_row + 1  # row after legend
    data_start_row = header_row + 1

    # Column layout for the comparison table
    #   A-D:  identity + plan link
    #   E-H:  rates (supply, usage/peak/offpeak)
    #   I-K:  solar FIT (tiers + details)
    #   L-M:  controlled load (rate + supply)
    #   N-O:  peak/offpeak %
    #   P:    usage cost
    #   Q:    solar credit
    #   R:    controlled load cost
    #   S:    net cost/day
    #   T:    net cost/month
    calc_columns = [
        ("Plan Name",                  "A"),
        ("Retailer",                   "B"),
        ("Tariff Type",                "C"),
        ("Plan URL",                   "D"),
        ("Supply (c/day)",             "E"),
        ("Usage Rate (c/kWh)",         "F"),
        ("Peak Rate (c/kWh)",          "G"),
        ("Off-Peak Rate (c/kWh)",      "H"),
        ("Solar FIT first tier (c/kWh)", "I"),
        ("Solar FIT thereafter (c/kWh)", "J"),
        ("Solar FIT Details",          "K"),
        ("CL Rate (c/kWh)",           "L"),
        ("CL Supply (c/day)",         "M"),
        ("Peak %",                     "N"),
        ("Off-Peak %",                 "O"),
        ("Usage Cost/day (c)",         "P"),
        ("Solar Credit/day (c)",       "Q"),
        ("CL Cost/day (c)",           "R"),
        ("Net Cost/day (c)",           "S"),
        ("Net Cost/month ($)",         "T"),
    ]

    # Write headers
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    for col_name, col_letter in calc_columns:
        cell = ws[f"{col_letter}{header_row}"]
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # ---- Write plan data rows with formulas ----
    data_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    cost_fill_even = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")  # green tint
    rate_fill_even = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")  # orange tint
    solar_fill_even = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")  # yellow tint
    id_fill_even = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")    # blue tint

    # Group columns for coloring
    id_cols = {"A", "B", "C", "D"}
    rate_cols = {"E", "F", "G", "H"}
    solar_cols = {"I", "J", "K"}
    cl_cols = {"L", "M", "R"}
    pct_cols = {"N", "O"}
    cost_cols = {"P", "Q", "S", "T"}
    cl_fill_even = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")  # purple tint

    # Sort plans by supply charge ascending as a neutral default
    sorted_plans = sorted(plans_data, key=lambda p: (p.get("Supply Charge (c/day)") or 999,
                                                      p.get("Usage Rate Min (c/kWh)") or 999))

    for row_offset, plan in enumerate(sorted_plans):
        row = data_start_row + row_offset
        is_tou = plan.get("Pricing Model") == "TOU"
        supply = plan.get("Supply Charge (c/day)") or 0
        # For SR plans: use Usage Rate Max (the most common rate / first block)
        usage_rate = plan.get("Usage Rate Max (c/kWh)") or 0
        peak_rate = plan.get("Peak Rate (c/kWh)") or 0
        offpeak_rate = plan.get("Off-Peak Rate (c/kWh)") or 0

        # Controlled load rates
        cl_rate = plan.get("CL Rate (c/kWh)") or 0
        cl_supply = plan.get("CL Supply (c/day)") or 0

        # Parse solar FIT tiers
        tiers = _parse_solar_fit_tiers(plan)
        if len(tiers) >= 2:
            # Tiered: first tier has a volume cap, second is the remainder
            fit_tier1_rate = tiers[0][0]
            fit_tier1_vol = tiers[0][1]
            fit_tier2_rate = tiers[1][0]
        elif len(tiers) == 1:
            # Flat FIT (single rate, no volume cap)
            fit_tier1_rate = tiers[0][0]
            fit_tier1_vol = 0  # no cap
            fit_tier2_rate = 0
        else:
            fit_tier1_rate = 0
            fit_tier1_vol = 0
            fit_tier2_rate = 0

        # A: Plan Name
        ws[f"A{row}"] = plan.get("Plan Name", "")
        ws[f"A{row}"].font = data_font

        # B: Retailer
        ws[f"B{row}"] = plan.get("Retailer", "")
        ws[f"B{row}"].font = data_font

        # C: Tariff Type
        ws[f"C{row}"] = plan.get("Tariff Type", "")
        ws[f"C{row}"].font = data_font

        # D: Plan URL (clickable hyperlink)
        plan_url = plan.get("Plan URL", "")
        if plan_url:
            ws[f"D{row}"].hyperlink = plan_url
            ws[f"D{row}"].value = "View Plan"
            ws[f"D{row}"].font = link_font
        else:
            ws[f"D{row}"] = ""
            ws[f"D{row}"].font = data_font

        # E: Supply (c/day) - static value
        ws[f"E{row}"] = supply
        ws[f"E{row}"].font = data_font
        ws[f"E{row}"].number_format = "0.00"

        # F: Usage Rate (c/kWh) - for SR plans
        ws[f"F{row}"] = usage_rate if not is_tou else ""
        ws[f"F{row}"].font = data_font
        if not is_tou:
            ws[f"F{row}"].number_format = "0.00"

        # G: Peak Rate (c/kWh) - for TOU plans
        ws[f"G{row}"] = peak_rate if is_tou else ""
        ws[f"G{row}"].font = data_font
        if is_tou:
            ws[f"G{row}"].number_format = "0.00"

        # H: Off-Peak Rate (c/kWh) - for TOU plans
        ws[f"H{row}"] = offpeak_rate if is_tou else ""
        ws[f"H{row}"].font = data_font
        if is_tou:
            ws[f"H{row}"].number_format = "0.00"

        # I: Solar FIT first tier rate (c/kWh)
        ws[f"I{row}"] = fit_tier1_rate
        ws[f"I{row}"].font = data_font
        ws[f"I{row}"].number_format = "0.00"

        # J: Solar FIT remainder tier rate (c/kWh)
        ws[f"J{row}"] = fit_tier2_rate
        ws[f"J{row}"].font = data_font
        ws[f"J{row}"].number_format = "0.00"

        # K: Solar FIT Details (text description)
        ws[f"K{row}"] = plan.get("Solar FIT Details", "")
        ws[f"K{row}"].font = data_font

        # L: CL Rate (c/kWh) - controlled load usage rate
        ws[f"L{row}"] = cl_rate
        ws[f"L{row}"].font = data_font
        ws[f"L{row}"].number_format = "0.00"

        # M: CL Supply (c/day) - controlled load daily supply charge
        ws[f"M{row}"] = cl_supply
        ws[f"M{row}"].font = data_font
        ws[f"M{row}"].number_format = "0.00"

        # N: Peak % (formula based on profile, only for TOU; 100% for SR)
        if is_tou:
            ws[f"N{row}"] = f"={peak_formula_inner}"
            ws[f"N{row}"].number_format = "0%"
        else:
            ws[f"N{row}"] = 1.0  # SR plans: all usage at the single rate
            ws[f"N{row}"].number_format = "0%"
        ws[f"N{row}"].font = data_font

        # O: Off-Peak % (formula based on profile, only for TOU; 0% for SR)
        if is_tou:
            ws[f"O{row}"] = f"={offpeak_formula_inner}"
            ws[f"O{row}"].number_format = "0%"
        else:
            ws[f"O{row}"] = 0.0
            ws[f"O{row}"].number_format = "0%"
        ws[f"O{row}"].font = data_font

        # P: Usage Cost/day (c)
        if is_tou:
            ws[f"P{row}"] = f"=$B$4*(N{row}*G{row}+O{row}*H{row})"
        else:
            ws[f"P{row}"] = f"=$B$4*F{row}"
        ws[f"P{row}"].font = data_font
        ws[f"P{row}"].number_format = "0.00"

        # Q: Solar Credit/day (c) - tiered calculation
        #    If tier1 has a volume cap:
        #      credit = MIN(export, cap) * tier1_rate + MAX(export - cap, 0) * tier2_rate
        #    If flat (no cap / cap=0):
        #      credit = export * tier1_rate
        if fit_tier1_vol > 0:
            # Tiered FIT
            ws[f"Q{row}"] = (
                f"=MIN($B$5,{fit_tier1_vol})*I{row}"
                f"+MAX($B$5-{fit_tier1_vol},0)*J{row}"
            )
        else:
            # Flat FIT (single rate)
            ws[f"Q{row}"] = f"=$B$5*I{row}"
        ws[f"Q{row}"].font = data_font
        ws[f"Q{row}"].number_format = "0.00"

        # R: CL Cost/day (c) - controlled load cost, only applied when B7 = "Yes"
        #    Formula: IF(B7="Yes", CL_usage * CL_rate + CL_supply, 0)
        ws[f"R{row}"] = f'=IF($B$7="Yes",$B$8*L{row}+M{row},0)'
        ws[f"R{row}"].font = data_font
        ws[f"R{row}"].number_format = "0.00"

        # S: Net Cost/day (c) = supply + usage - solar + controlled load
        ws[f"S{row}"] = f"=E{row}+P{row}-Q{row}+R{row}"
        ws[f"S{row}"].font = data_font
        ws[f"S{row}"].number_format = "0.00"

        # T: Net Cost/month ($) = net_cost_day * 30.44 / 100
        ws[f"T{row}"] = f"=S{row}*30.44/100"
        ws[f"T{row}"].font = data_font
        ws[f"T{row}"].number_format = "$#,##0.00"

        # Apply borders and alternating row coloring
        for col_name, col_letter in calc_columns:
            cell = ws[f"{col_letter}{row}"]
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if row % 2 == 0:
                if col_letter in id_cols:
                    cell.fill = id_fill_even
                elif col_letter in rate_cols:
                    cell.fill = rate_fill_even
                elif col_letter in solar_cols:
                    cell.fill = solar_fill_even
                elif col_letter in cl_cols:
                    cell.fill = cl_fill_even
                elif col_letter in cost_cols:
                    cell.fill = cost_fill_even

    # ---- Auto-fit column widths ----
    col_widths = {
        "A": 40, "B": 20, "C": 12, "D": 12, "E": 15,
        "F": 18, "G": 18, "H": 20, "I": 22, "J": 24,
        "K": 30, "L": 18, "M": 18,
        "N": 10, "O": 12,
        "P": 20, "Q": 20, "R": 18, "S": 18, "T": 20,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze above the data table header
    ws.freeze_panes = f"A{data_start_row}"

    # Add auto-filter on the data table
    last_row = data_start_row + len(sorted_plans) - 1
    if last_row >= data_start_row:
        ws.auto_filter.ref = f"A{header_row}:T{last_row}"

    return header_row, len(sorted_plans)


# ---------------------------------------------------------------------------
# VBA Macro Injection
# ---------------------------------------------------------------------------


def _inject_vba_and_save_as_xlsm(xlsx_path: str, header_row: int, plan_count: int) -> str:
    """Open the .xlsx with Excel via COM, inject a VBA sort macro, and re-save as .xlsm.

    The macro fires on Worksheet_Change for the Plan Calculator sheet and
    re-sorts the plan data table by column Q (Net Cost/month) ascending
    whenever cells B4, B5, or B6 are modified.

    Args:
        xlsx_path: Path to the .xlsx file produced by openpyxl.
        header_row: The row number containing column headers in Plan Calculator.
        plan_count: Number of plan data rows.

    Returns:
        Path to the new .xlsm file. The original .xlsx is deleted.
    """
    import pythoncom
    import win32com.client as win32

    data_start = header_row + 1
    data_end = header_row + plan_count

    vba_code = f'''
Private Sub Worksheet_Change(ByVal Target As Range)
    ' Auto-sort plan table by Net Cost/month (column T) when inputs change
    Dim watchRange As Range
    Set watchRange = Me.Range("B4:B8")

    If Not Intersect(Target, watchRange) Is Nothing Then
        Application.EnableEvents = False
        On Error GoTo Cleanup

        Dim sortRange As Range
        Set sortRange = Me.Range("A{data_start}:T{data_end}")

        sortRange.Sort _
            Key1:=Me.Range("T{data_start}"), _
            Order1:=xlAscending, _
            Header:=xlNo, _
            OrderCustom:=1, _
            MatchCase:=False, _
            Orientation:=xlTopToBottom, _
            DataOption1:=xlSortNormal

Cleanup:
        Application.EnableEvents = True
    End If
End Sub
'''

    abs_xlsx = str(Path(xlsx_path).resolve())
    xlsm_path = abs_xlsx.replace(".xlsx", ".xlsm")

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(abs_xlsx)

        # Find the "Plan Calculator" sheet
        calc_sheet = None
        for ws in wb.Worksheets:
            if ws.Name == "Plan Calculator":
                calc_sheet = ws
                break

        if calc_sheet is not None:
            # Access the VBA project and inject the macro into the sheet's code module
            vba_module = calc_sheet.CodeModule
            vba_module.DeleteLines(1, vba_module.CountOfLines)
            vba_module.AddFromString(vba_code)

        # Save as macro-enabled workbook (.xlsm) - file format 52
        wb.SaveAs(xlsm_path, FileFormat=52)
        wb.Close(SaveChanges=False)
    finally:
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()

    # Remove the intermediate .xlsx
    try:
        os.remove(abs_xlsx)
    except OSError:
        pass

    return xlsm_path


# ---------------------------------------------------------------------------
# Distributor Selection Logic
# ---------------------------------------------------------------------------


def select_distributor_interactive(postcode: str, fuel_type: str,
                                   customer_type: str) -> list[dict]:
    """Interactively select distributor(s) for a multi-distributor postcode.

    Returns a list of dicts: [{"id": "13", "name": "Endeavour", "plan_count": 1176}, ...]
    representing the distributor(s) the user wants to fetch.
    """
    print("\n  Checking available electricity distributors...")
    distributors = fetch_distributors(postcode, fuel_type)

    if not distributors:
        print("  No distributors found for this postcode.")
        return []

    if len(distributors) == 1:
        print(f"  Single distributor: {distributors[0]['name']} (ID: {distributors[0]['id']})")
        return distributors

    # Multiple distributors - probe each to see which ones have plans
    print(f"  Found {len(distributors)} distributors. Checking plan availability...")
    available = []
    for d in distributors:
        count = probe_distributor_plans(postcode, d["id"], fuel_type, customer_type)
        if count > 0:
            d["plan_count"] = count
            available.append(d)
            print(f"    [{len(available)}] {d['name']} (ID: {d['id']}) - {count} plans")
        elif count == 0:
            print(f"    [ ] {d['name']} (ID: {d['id']}) - no plans (skipped)")
        else:
            print(f"    [?] {d['name']} (ID: {d['id']}) - error checking (skipped)")

    if not available:
        print("  No distributors returned plans for this postcode. Try the basic scraper.")
        return []

    if len(available) == 1:
        print(f"\n  Only one distributor has plans: {available[0]['name']}")
        return available

    # Ask user to choose
    print(f"\n  This postcode has {len(available)} active distributors.")
    print("  Your electricity distributor depends on your street address.")
    print("  Check your electricity bill or contact your provider if unsure.")
    print()
    print("  Select a distributor:")
    for i, d in enumerate(available, 1):
        print(f"    [{i}] {d['name']} ({d['plan_count']} plans)")
    print(f"    [A] Fetch ALL distributors ({sum(d['plan_count'] for d in available)} plans total)")
    print()

    while True:
        choice = input("  Enter your choice (number or A): ").strip().upper()
        if choice == "A":
            return available
        try:
            idx = int(choice) - 1
            if 0 <= idx < len(available):
                return [available[idx]]
        except ValueError:
            pass
        print(f"  Invalid choice. Enter 1-{len(available)} or A.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Enhanced scraper for Energy Made Easy with distributor selection"
    )
    parser.add_argument(
        "postcode",
        nargs="?",
        help="Australian postcode to search (e.g. 2000, 2850, 4000)",
    )
    parser.add_argument(
        "--fuel",
        choices=["electricity", "gas"],
        default="electricity",
        help="Fuel type to compare (default: electricity)",
    )
    parser.add_argument(
        "--type",
        dest="customer_type",
        choices=["residential", "business"],
        default="residential",
        help="Customer type (default: residential)",
    )
    parser.add_argument(
        "--dist",
        dest="distributor",
        default=None,
        help='Distributor ID (e.g. 13 for Endeavour) or "all" to fetch every distributor',
    )
    parser.add_argument(
        "--controlled-load",
        dest="controlled_load",
        action="store_true",
        default=False,
        help="Include plans with controlled load (hot water / pool pump circuits)",
    )
    parser.add_argument(
        "--no-filter",
        dest="no_filter",
        action="store_true",
        default=False,
        help="Disable all plan filtering (get ALL plans from the API)",
    )
    args = parser.parse_args()

    postcode = args.postcode
    if not postcode:
        postcode = input("Enter an Australian postcode (e.g. 2000): ").strip()

    if not postcode or not postcode.isdigit() or len(postcode) != 4:
        print("Error: Please enter a valid 4-digit Australian postcode.")
        sys.exit(1)

    fuel_type = "E" if args.fuel == "electricity" else "G"
    customer_type = "R" if args.customer_type == "residential" else "B"

    include_controlled_load = args.controlled_load
    no_filter = args.no_filter

    print(f"\n{'='*60}")
    print(f"  Energy Made Easy Enhanced Scraper")
    print(f"{'='*60}")
    print(f"  Postcode:      {postcode}")
    print(f"  Fuel Type:     {args.fuel.title()}")
    print(f"  Customer Type: {args.customer_type.title()}")
    if args.distributor:
        print(f"  Distributor:   {args.distributor}")
    if no_filter:
        print(f"  Filtering:     DISABLED (showing all plans)")
    elif include_controlled_load:
        print(f"  Filtering:     Including controlled load plans")
    print(f"{'='*60}\n")

    # Step 1: Validate postcode
    print("Step 1: Validating postcode...")
    try:
        locations = validate_postcode(postcode)
    except requests.exceptions.RequestException as e:
        print(f"Error validating postcode: {e}")
        sys.exit(1)

    if not locations:
        print(f"Error: No locations found for postcode {postcode}.")
        print("Note: Energy Made Easy covers NSW, QLD, SA, TAS, and ACT only.")
        sys.exit(1)

    location_names = [f"{loc['location'].title()}, {loc['state']}" for loc in locations]
    print(f"  Found locations: {', '.join(location_names[:5])}", end="")
    if len(location_names) > 5:
        print(f" and {len(location_names) - 5} more")
    else:
        print()

    # Step 2: Determine distributor(s)
    print("\nStep 2: Determining electricity distributor(s)...")

    if args.distributor and args.distributor.lower() == "all":
        # Fetch all distributors
        distributors = fetch_distributors(postcode, fuel_type)
        selected = []
        for d in distributors:
            count = probe_distributor_plans(postcode, d["id"], fuel_type, customer_type)
            if count > 0:
                d["plan_count"] = count
                selected.append(d)
                print(f"  {d['name']} (ID: {d['id']}) - {count} plans")
        if not selected:
            print("  No distributors returned plans. Trying without distributor filter...")
            selected = [{"id": "", "name": "Auto", "plan_count": 0}]
    elif args.distributor:
        # Use the specified distributor ID
        dist_id = args.distributor
        distributors = fetch_distributors(postcode, fuel_type)
        dist_name = next((d["name"] for d in distributors if d["id"] == dist_id), f"ID {dist_id}")
        selected = [{"id": dist_id, "name": dist_name, "plan_count": 0}]
        print(f"  Using specified distributor: {dist_name} (ID: {dist_id})")
    else:
        # Interactive selection
        selected = select_distributor_interactive(postcode, fuel_type, customer_type)

    if not selected:
        print("Error: No distributor selected or available.")
        sys.exit(1)

    # Step 3: Fetch plans for each selected distributor
    print("\nStep 3: Fetching energy plans...")
    all_raw_plans = []
    distributor_names = []
    start_time = time.time()

    for dist in selected:
        dist_id = dist["id"]
        dist_name = dist["name"]
        distributor_names.append(dist_name)

        print(f"  Fetching plans for {dist_name}...")
        try:
            raw_plans = fetch_plans(postcode, fuel_type, customer_type, dist_id)
        except requests.exceptions.RequestException as e:
            print(f"  Error fetching plans for {dist_name}: {e}")
            continue

        print(f"    Found {len(raw_plans)} plans from API")
        for plan in raw_plans:
            plan["_distributor_name"] = dist_name
        all_raw_plans.extend(raw_plans)

    # Step 3b: Filter plans
    if no_filter:
        print(f"\n  Filtering disabled: keeping all {len(all_raw_plans)} plans")
        filtered_plans = all_raw_plans
    else:
        # Interactive: ask about controlled load if not specified via CLI
        if not include_controlled_load and sys.stdin.isatty():
            print()
            print("  Do you have a controlled load circuit (e.g. off-peak hot water,")
            print("  pool pump, or floor heating on a separate meter/tariff)?")
            cl_answer = input("  Enter Y or N [N]: ").strip().upper()
            if cl_answer == "Y":
                include_controlled_load = True
                print("  Including controlled load plans.")
            else:
                print("  Excluding controlled load plans.")

        filtered_plans, filter_stats = filter_plans(
            all_raw_plans,
            include_controlled_load=include_controlled_load,
            include_demand=False,
        )

        print(f"\n  Filtering results:")
        print(f"    Total from API:               {filter_stats['total']}")
        if filter_stats["demand_filtered"] > 0:
            print(f"    Removed (demand charge):       {filter_stats['demand_filtered']}  (require demand meter)")
        if filter_stats["controlled_load_filtered"] > 0:
            print(f"    Removed (controlled load):     {filter_stats['controlled_load_filtered']}  (require CL circuit)")
        print(f"    Plans available to you:        {filter_stats['kept']}")

    # Step 4: Process filtered plans
    print(f"\nStep 4: Processing {len(filtered_plans)} plans...")
    all_plans_data = []

    errors = 0
    for plan in filtered_plans:
        dist_name = plan.get("_distributor_name", "N/A")
        try:
            row = process_plan(plan, postcode, dist_name)
            all_plans_data.append(row)
        except Exception as e:
            errors += 1
            plan_name = plan.get("planData", {}).get("planName", "Unknown")
            print(f"    Warning: Failed to process plan '{plan_name}': {e}")

    if errors:
        print(f"    {errors} plans had processing errors")

    elapsed = time.time() - start_time

    if not all_plans_data:
        print(f"\nNo plans found for postcode {postcode}.")
        sys.exit(1)

    print(f"\n  Total: {len(all_plans_data)} plans processed in {elapsed:.1f}s")

    # Step 5: Export to Excel
    print("\nStep 5: Exporting to Excel spreadsheet...")
    distributor_info = " + ".join(distributor_names) if len(distributor_names) <= 3 else f"{len(distributor_names)} distributors"
    try:
        filepath = export_to_excel(all_plans_data, postcode, fuel_type, customer_type, distributor_info)
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        sys.exit(1)

    print(f"  Saved to: {filepath}")

    # Open the file automatically
    abs_path = str(Path(filepath).resolve())
    print(f"  Opening {abs_path}...")
    import platform
    import subprocess
    if platform.system() == "Windows":
        os.startfile(abs_path)
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", abs_path])
    else:
        subprocess.Popen(["xdg-open", abs_path])

    # Summary stats
    print(f"\n{'='*60}")
    print(f"  COMPLETE - Summary")
    print(f"{'='*60}")
    print(f"  Total plans:        {len(all_plans_data)}")
    print(f"  Distributor(s):     {distributor_info}")

    retailers = set(p["Retailer"] for p in all_plans_data)
    print(f"  Unique retailers:   {len(retailers)}")

    sr_count = sum(1 for p in all_plans_data if p["Pricing Model"] == "SR")
    tou_count = sum(1 for p in all_plans_data if p["Pricing Model"] == "TOU")
    print(f"  Single rate plans:  {sr_count}")
    print(f"  Time of use plans:  {tou_count}")

    solar_count = sum(1 for p in all_plans_data if (p.get("Solar FIT Max (c/kWh)") or 0) > 0)
    print(f"  Plans with solar:   {solar_count}")

    # Show cheapest plans
    cheap = [p for p in all_plans_data if p.get("Est. Cost/Year (Medium Usage)") is not None]
    if cheap:
        cheap.sort(key=lambda p: p["Est. Cost/Year (Medium Usage)"])
        print(f"\n  Top 5 Cheapest (Medium Usage, with discounts):")
        for i, p in enumerate(cheap[:5], 1):
            dist_label = f" [{p['Distributor']}]" if len(selected) > 1 else ""
            print(f"    {i}. {p['Plan Name']} ({p['Retailer']}{dist_label}) - ${p['Est. Cost/Year (Medium Usage)']:,}/yr")

    # Show best solar FIT
    solar_with_fit = [p for p in all_plans_data if (p.get("Solar FIT Max (c/kWh)") or 0) > 0]
    if solar_with_fit:
        solar_with_fit.sort(key=lambda p: p.get("Solar FIT Max (c/kWh)") or 0, reverse=True)
        print(f"\n  Top 5 Best Solar Feed-in Tariffs:")
        for i, p in enumerate(solar_with_fit[:5], 1):
            fit_max = p["Solar FIT Max (c/kWh)"]
            fit_min = p.get("Solar FIT Min (c/kWh)", fit_max)
            if fit_min == fit_max:
                fit_str = f"{fit_max}c/kWh"
            else:
                fit_str = f"{fit_min} - {fit_max}c/kWh"
            dist_label = f" [{p['Distributor']}]" if len(selected) > 1 else ""
            print(f"    {i}. {p['Plan Name']} ({p['Retailer']}{dist_label}) - {fit_str}")

    print(f"\n  Spreadsheet saved to: {filepath}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
